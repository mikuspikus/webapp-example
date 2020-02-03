from django.shortcuts import render
from django.views import View
from django.http import HttpRequest, HttpResponseRedirect, HttpResponse
from django.template import loader, Context

from typing import Union

import openpyxl as xl
import datetime

from . import models
from . import forms 

class SuccessView(View):
    template = 'application/success.html'

    def get(self, request: HttpRequest, *args, **kwargs) -> HttpResponse:
        return render(request = request, template_name = self.template)

class IndexView(View):
    redirect_to = 'success'
    form_template = 'application/index.html'
    model = models.Application
    form = forms.ApplicationForm

    def get(self, request: HttpRequest, *args, **kwargs) -> HttpResponse:
        form = self.form()

        return render(
            request = request,
            template_name = self.form_template,
            context = {'form': form}
        )

    def post(self, request: HttpRequest, *args, **kwargs) -> Union[HttpResponse, HttpResponseRedirect]:
        form = self.form(data = request.POST)

        is_valid = form.is_valid()
        if is_valid:
            form.save()
            return HttpResponseRedirect(
                redirect_to = self.redirect_to
            )

        return render(
            request = request,
            status = 400,
            template_name = self.form_template,
            context = {'form': form}
        )

class UploadView(View):
    filename_template = 'applications-{date}.xlsx'
    model = models.Application
    title = 'Applications'
    header = ('name', 'surname', 'phone', 'applied_at')

    def get(self, request: HttpRequest, *args, **kwargs) -> HttpResponse:
        response = HttpResponse(content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = self.filename_template.format(date = datetime.datetime.now().strftime('%Y-%m-%d'))
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        workbook = xl.Workbook()
        worksheet = workbook.active
        worksheet.title = self.title

        r_num = 1
        for c_num, c_header in enumerate(self.header, 1):
            cell = worksheet.cell(row = r_num, column = c_num, value = c_header)

        data = [row.row() for row in self.model.objects.all()]

        for data_row in data:
            r_num += 1
            for c_num, value in enumerate(data_row, 1):
                cell = worksheet.cell(row = r_num, column = c_num, value = value)

        workbook.save(response)
        return response

